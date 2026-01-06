import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os


def league_data(some_id, season):
    url = "https://v3.football.api-sports.io/fixtures"

    query_string = {
        "league": some_id,         
        "season": season        
    }

    headers = {
    'x-rapidapi-key': 'eb12aee9fb1f5d338a47bf720c4a9dd5',
    'x-rapidapi-host': 'v3.football.api-sports.io'
    }

    response = requests.get(url, headers=headers, params=query_string)
    data = response.json()

    print("status:", response.status_code)
    print("errors:", data.get("errors"))
    print("message:", data.get("message"))
    print("results:", data.get("results"))
    print("paging:", data.get("paging"))

    match = data['response']

    print(f"Season: {season}")
    print(f"Number of matches returned: {len(match)}")
    print(f"Full API response (first 2 items): {match[:2]}")

    df = pd.DataFrame([{
        'date': m['fixture']['date'],
        'id' : m['fixture']['id'],
        'referee' : m['fixture']['referee'],
        'home_team' : m['teams']['home']['name'],
        'away_team' : m['teams']['away']['name'],
        'home_goals' : m['goals']['home'],
        'away_goals' : m['goals']['away'],
    }for m in match])

    df['result'] = df.apply(
        lambda row: 'home' if row['home_goals'] > row['away_goals'] 
        else 'away' if row['home_goals'] < row['away_goals'] 
        else 'draw',
        axis=1
    )   
    
    assert df.shape[0] > 0
    assert df["date"].notna().all()
    assert df["home_goals"].notna().all()
    assert df["away_goals"].notna().all()

    df["date"] = pd.to_datetime(df["date"], utc=True).dt.tz_convert(None)
    df = df.sort_values("date").drop_duplicates(subset=["id"])

    league_name = match[0]['league']['name']

    df["total_goals"] = df["home_goals"] + df["away_goals"]
    df["date"] = pd.to_datetime(df["date"])
    df["goal_diff"] = df["home_goals"] - df["away_goals"]
    
    
    output_dir = "D:/output_England"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_path = f"D:/output_England/{league_name}-{season}.csv"
    df.to_csv(output_path, index=False)
    print("Saving to:", output_path)
    return output_path

def file_adjust(output_path):
    wb = load_workbook(output_path)
    ws= wb.active

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_path)



#league_id = [39, 79, 61, 140]
    """
    for league in league_id:
    for i in range(2022, 2024):
        output_path = league_data(league,i)
    """
if __name__ == "__main__":
    league_id = [39]
    for league in league_id:
        for season in range(2022, 2024):  # 2022, 2023
            league_data(league, season)
